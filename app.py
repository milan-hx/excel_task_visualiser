import json
import pandas as pd
from openpyxl import Workbook, styles
import datetime

# Function to return the date object for a given date string
def get_date(date_string):
    date = datetime.datetime.strptime(date_string, '%m/%d/%Y')
    return date

# Function to load the config file and return it in a json object
def load_config():
    try:
        with open("config.json", "r") as f:
            config = json.load(f)
    except:
        print("Error loading config file")
        exit()

    return config

# Function to Load and format input data and return it in a json object
def get_input_json(filepath):
    df = pd.read_excel(filepath, header=None)
    columns = df.iloc[2].tolist()
    columns[0] = "task"
    df.columns = columns
    df = df.iloc[3:]
    df = df.reset_index(drop=True)
    return json.loads(df.to_json(orient="records"))

# Function to return the fill object for a given color
def get_fill(color):
    my_color = styles.colors.Color(rgb='FFFFFF')
    if color == "blue":
        my_color = styles.colors.Color(rgb='ADD8E6')
    elif color == "green":
        my_color = styles.colors.Color(rgb='00FF00')

    return styles.fills.PatternFill(patternType='solid', fgColor=my_color)

# Function to return true if the cell is in the range of the project stage dates
def in_range(c_start, c_end, d_start, d_end, prod_or_need):
    fill_cell = False
    if prod_or_need:
        if d_end <= c_end and d_end >= c_start:
            fill_cell = True
    else:
        if d_start >= c_start and d_start <= c_end:
            fill_cell = True
        elif d_start <= c_start and d_end >= c_start:
            fill_cell = True
        elif d_end <= c_end and d_end >= c_start:
            fill_cell = True
    return fill_cell



# Function to set the cell colors for a task record based on the project stage dates
def fill_cells(stage, fill_config, record, record_row):
    prod_or_need = False
    if stage == "dev":
        s_start = get_date(record["DevBegin"])
        s_end = get_date(record["DevEnd"])
    elif stage == "test":
        s_start = get_date(record["TestBeg"])
        s_end = get_date(record["TestEnd"])
    elif stage == "prod":
        prod_or_need = True
        s_start = None
        s_end = get_date(record["*"])
    elif stage == "need_by":
        prod_or_need = True
        s_start = None
        s_end = get_date(record["#"])
    ws[f"A{record_row}"].value = record["task"]

    for cell in letter_indexing[1:]:
        if ws[f"{cell}1"].value == None:
            continue
        cell_start = datetime.datetime.strptime(
            ws[f"{cell}1"].value.split("-")[0], '%m/%d/%Y')
        cell_end = datetime.datetime.strptime(
            ws[f"{cell}1"].value.split("-")[1], '%m/%d/%Y')

        if in_range(cell_start, cell_end, s_start, s_end, prod_or_need):
            if fill_config["use_symbol"]:
                ws[f"{cell}{record_row}"].value = fill_config["symbol"]
            else:
                ws[f"{cell}{record_row}"].fill = get_fill(fill_config["color"])

# Init openpyxl workbook
wb = Workbook()
ws = wb.active

# Init configuration
record_row = 2
letter_indexing = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
config = load_config()
stage_config = config["stage_config"]
enclosing_date_range = config["enclosing_date_range"]

# Load and format input data
json_data = get_input_json(config["input_filepath"])


# Fill date range cells based on enclosing date range
start_date = get_date(enclosing_date_range["start"])
end_date = get_date(enclosing_date_range["end"])
temp_date = start_date
next_date = temp_date
index_count = 1
while temp_date < end_date:
    next_date = temp_date + datetime.timedelta(days=13)
    ws[f"{letter_indexing[index_count]}1"] = f"{temp_date.month}/{temp_date.day}/{temp_date.year}-{next_date.month}/{next_date.day}/{next_date.year}"
    temp_date = next_date + datetime.timedelta(days=1)
    index_count += 1

# Fill cells for each task record
stage_list = stage_config.keys()
for record in json_data:
    for stage in stage_list:
        fill_cells(stage, stage_config[stage], record, record_row)
    record_row += 1

#Save workbook to output file
wb.save(config["output_filepath"])
